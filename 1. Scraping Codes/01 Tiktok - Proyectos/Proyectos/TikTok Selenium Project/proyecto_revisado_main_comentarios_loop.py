#MODULOS DE SELENIUM
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.utils.dataframe import dataframe_to_rows

#MODULOS SECUNDARIOS
import pandas as pd
import numpy as np
import os
import re
import time
import ssl
import sys
import math
import traceback
import openpyxl
import winsound


#MODULO ALERTAS DESACTIVADAS
import warnings
warnings.filterwarnings('ignore')

#MODULO PROPIOS
from funciones_primarias import validacion, comments_disponibles
from funciones_secundarias import barra_carga,scroll_down

'''Protocolos SSL'''
ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

def scrap_tiktok(url):

  '''-------------------------------------------------------------------------------------------------------'''

  '''
  OBJETO DRIVER 
  - OPCIONES 
  - URL 
  '''

  #Opciones Firefox
  firefox_profile = webdriver.FirefoxOptions()
  firefox_profile.set_preference("browser.download.folderList", 2)
  firefox_profile.set_preference("browser.download.dir", 'descargas')
  firefox_profile.set_preference("browser.helperApps.neverAsk.saveToDisk", "application/pdf")

  #Navegador Firefox con el perfil configurado
  driver = webdriver.Firefox(options=firefox_profile)


  #Instancia driver -  Acceso a la URL
  PAGINA_URL = url
  driver.get(PAGINA_URL)


  '''-------------------------------------------------------------------------------------------------------'''

  ''' 
  VALIDACION 
    - Funcionamiento de acceso a la URL, navegacion y captura de segmentos del en X.Path
    - Prueba del Scroll y de la Estrategia de Espera para la carga de mensajes
  '''
  seccion_comments = '/html/body/div[1]/div[2]/div[2]/div/div[2]/div[1]/div[2]/div[2]/div/div['

  #Ejecucion de las pruebas de validacion, comentarlas cuando se aplique la aplicacion general

  barra_carga(20)
  scroll_down(5,1,1080,driver)
  validacion(driver,seccion_comments)

  '''-------------------------------------------------------------------------------------------------------'''


  '''PROGRAMA WEB SCRAPING TIK-TOK'''

  #VARIABLES PARA EL SCRAPING
  comentarios_recorridos = list()
  centinela = True
  comment_externo = 1
  # comment_interno = 1 
  base_final = pd.DataFrame()


  #BOT - CODIGO DE EJECUCION GENERAL

  '''
  El codigo tiene como punto de referencia el container del comentario, accede a los datos del usuario por medio de un loop,
  en el cual se extraen los datos basicos, incluidos el numero de replies y likes, para posteriormente 
  ingresar al contenedor de las respuestas de su comentario. Repitiendo un la misma logida de recuperacion de datos que el proceso anterior
  '''

  while centinela == True:
    try:
      ##Validacion de acceso a la URL
      try:
        driver.find_element(By.XPATH, value= '/html/body/div[1]/div[2]/div[2]/div/div[2]/div[1]/div[2]/div[2]/div/div[1]')
        print('INICIO_Usuario: ')
        time.sleep(5)
      except:
        pass

      ##NAVEGACION Y DATOS PRINCIPALES    
      # Level 1 comment principal
      level = 1
      # Detectar el display name del usuario (i)
      try:
        display_name = driver.find_element(By.XPATH, value = seccion_comments + str(comment_externo) + "]/div[1]/div[1]/a/span").text
        print(display_name)
      except:
        centinela = False
        break
      # 01 Detectar nombre de usuario
      username_thread = driver.find_element(by=By.PARTIAL_LINK_TEXT,value=display_name).get_attribute("href")
      # 02 Detectar comment principal
      comment = driver.find_element(By.XPATH, value = seccion_comments + str(comment_externo) + "]/div[1]/div[1]/p[1]/span").text
      # 03 Detectar fecha
      fecha = driver.find_element(By.XPATH, value = seccion_comments + str(comment_externo) + "]/div[1]/div[1]/p[2]/span[1]").text
      # 04 Detectar número de likes
      likes = driver.find_element(By.XPATH, value = seccion_comments + str(comment_externo) + "]/div[1]/div[1]/p[2]/div/span").text
      # 05 Detectar el número de respuestas al comentario (answers)
      try:
        
        if comment_externo != 1:
          answers = driver.find_element(By.XPATH, value = seccion_comments + str(comment_externo) + "]/div[2]/div/p").text
          result = int(''.join([char for char in answers if char.isdigit()]))

          print(f'Numero de Respuestas {answers} {username_thread}')

        else: 
          answers = driver.find_element(By.XPATH, value = seccion_comments + str(comment_externo) + "]/div[2]/div[2]/p").text
          result = int(''.join([char for char in answers if char.isdigit()]))

          print(f'Numero de Respuestas {answers} {username_thread}')

      except:
        traceback.print_exc
        try:
          scroll_down(3, 1, 1030)
          answers = driver.find_element(By.XPATH, value = seccion_comments + str(comment_externo) + "]/div[2]/div/[2]p").text
          # result = answers[answers.find('(')+1:answers.find(')')]
          result = int(''.join([char for char in answers if char.isdigit()]))
          print(f'Numero de Respuestas Excepcional {answers}')
          
        except:
          answers = 0
          result = 0

      answers = result
      print(f'Validacion Respuestas {answers}')
      
      d = {"link":PAGINA_URL,"level": level, "display_name": display_name, "username":username_thread, "comment":comment, "fecha": fecha, "likes":likes, "replies":answers, "Thread_author": username_thread}
      
      base_temporal = pd.DataFrame(data=d,index=[comment_externo])
      base_temporal = base_temporal.rename(columns=base_temporal.iloc[0]).drop(base_temporal.index[0])
      # base_final = pd.concat([base_final, base_temporal], ignore_index=True)

      # VALIDACIONES DATOS BASICOS 

      # print(base_final)
      
      comentarios_recorridos.append(comment_externo)

      comment_externo = comment_externo + 1
      
  ##################################################################################################################

      ##  REPORTE POR PANTALLA DEL PROGRESO DEL BOT
      print("Acabé el comment " + str(comment_externo - 1) + ' perteneciente al usuario: ' + str(display_name), base_temporal)


      ##  CARGA DE LA LINEA CON LOS DATOS RECIEN COLECTADOS AL EXCEL DE FORMA CONTINUA

      path_archivo = 'base_columbia_tiktok_prueba_espa.xlsx'

      if not os.path.exists(path_archivo):
          workbook = openpyxl.Workbook()
          ws = workbook.create_sheet('data_consolidada')
          ws.append(["link", "level", "display_name", "username", "comment", "fecha", "likes", "replies", "Thread_author"])
          workbook.active = ws
          workbook.save(path_archivo)

      workbook = openpyxl.load_workbook('base_columbia_tiktok_prueba_espa.xlsx')
      workbook.active = workbook['data_consolidada']
      ws = workbook.active

      for r in dataframe_to_rows(base_temporal, index=False, header=True):
          ws.append(r)

      workbook.save('base_columbia_tiktok_prueba_espa.xlsx')

      # base_final.to_excel('base_columbia_prueba_tiktok.xlsx')

      ##########################################################

      if comment_externo % 15 == 1:
        scroll_down(10, 1, 1080,driver)
        try:
          driver.find_element(By.XPATH, value= '/html/body/div[7]/div/div[1]/div[2]/div')
          print('Hay captcha...')

          winsound.Beep(1000, 2000) 

          time.sleep(20)
        except:
          pass

    except Exception as e:

      
      # time.sleep(5)
      # scroll_down(3, 1, 850,driver)

      # with open('errores_progreso.txt', 'w') as fp:
      #   fp.write(str(comentarios_recorridos[-1]))

      print(f'Error {type(e)}', e)

      time.sleep(10)
      # base_final.to_excel('base_columbia_tiktok_prueba.xlsx')


  driver.close()
  # base_final.to_excel('base_columbia_tiktok_prueba.xlsx')


urls = [

'https://vm.tiktok.com/ZMMufR2a4/',
'https://vm.tiktok.com/ZMMuytpga/',
'https://www.tiktok.com/@pichachela/video/7142157393223535877',
'https://vm.tiktok.com/ZMMuycNLq/',
'https://vm.tiktok.com/ZMMuyTd8p/',
'https://www.tiktok.com/@darien_yatsugare/video/7192055523087289605',

]

# Iterar sobre la lista de URLs y llamar a la función para cada una
for url in urls:
    scrap_tiktok(url)